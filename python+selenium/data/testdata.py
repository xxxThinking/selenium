'''
@File:testdata.py
@DateTime:2021/12/10 17:14
@Author:sweet
@Desc:
'''

import openpyxl

class ReadWrite:
    def __init__(self,file,num):
        self.file = file
        self.num = num

    def read(self):
        wb = openpyxl.load_workbook(self.file)
        sheets = wb.sheetnames
        table = wb[sheets[self.num-1]]
        table = wb.active
        row = table.max_row
        col = table.max_column
        list1 = []
        for i in range(2, row + 1):
            list2 = []
            for j in range(1, col + 1):
                values = table.cell(i,j).value
                list2.append(values)
            list1.append(list2)
        return  list1