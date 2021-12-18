'''
@File:testdata.py
@DateTime:2021/12/10 17:14
@Author:sweet
@Desc:
'''

import openpyxl
from selenium_day05.config.config_1 import file,sheet

class ReadWrite:
    def __init__(self,file,sheet):
        self.file = file
        self.sheet = sheet

    def read(self):
        wb = openpyxl.load_workbook(self.file)
        table = wb[self.sheet]
        table = wb.active
        row = table.max_row
        col = table.max_column
        list1 = []
        for i in range(2, row + 1):
            list2 = []
            for j in range(1, col + 1):
                values = table.cell(i,j).value
                list2.append(values)
            list1.append(list2)
        return  list1
if __name__=="__main__":
    data = ReadWrite(file,sheet)
    print(data.read())